from openpyxl import load_workbook
from openpyxl import Workbook
import os
def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for data in _DATA:
            sheet.append(data)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        sheet.column_dimensions['F'].width = 40
        sheet.column_dimensions['G'].width = 40
        sheet.column_dimensions['H'].width = 40
        book.save(_FILENAME)

FILENAME = "디테.xlsx"
headerList = ['상품URL', '상품명','소비자가', '할인가','상품코드', '옵션', '대표이미지', '본문이미지']
save_excel(FILENAME,None,headerList)
save_excel(FILENAME,dataList,None)