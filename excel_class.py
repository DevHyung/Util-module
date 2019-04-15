"""
Copyright 2019 DevHyung
   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at
       http://www.apache.org/licenses/LICENSE-2.0
   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
"""
"""
This file is manual version.
Auto version To Be Announced 
"""
from openpyxl import load_workbook
from openpyxl import Workbook
import os

class ExcelDriver:
    def __init__(self, _fileName,_header):
        self.fileName = _fileName + ".xlsx"
        self.header = _header
        self.create_File()
    def create_File(self):
        if  os.path.exists(self.fileName):
            self.log('i','exist file name ')
        else:  # 새로만드는건
            book = Workbook()
            sheet = book.active
            sheet.title = 'default'
            sheet.append(self.header)
            # if modify cell width, write down
            #sheet.column_dimensions['A'].width = 40
            book.save(self.fileName)

    def append_data(self,_data):
        '''

        :param _data: 1-d list type input data to excel ,
        :return:
        '''
        book = load_workbook(self.fileName)
        sheet = book.active
        sheet.append(_data)
        book.save(self.fileName)

    def append_data_list(self,_dataList):
        '''
        :param _dataList: 2-d list type input data to excel ,
        :return:
        '''
        book = load_workbook(self.fileName)
        sheet = book.active
        for data in _dataList:
            sheet.append(data)
        book.save(self.fileName)

    @staticmethod
    def log(tag, text):
        # Info tag
        if (tag == 'i'):
            print("[INFO] " + text)
        # Error tag
        elif (tag == 'e'):
            print("[ERROR] " + text)
        # Success tag
        elif (tag == 's'):
            print("[SUCCESS] " + text)
if __name__ == "__main__":
    FILENAME = "인스타그램" # without file extention
    headerList = ['상품URL', '상품명', '소비자가', '할인가', '상품코드', '옵션', '대표이미지', '본문이미지']
    excel = ExcelDriver(FILENAME,headerList)

